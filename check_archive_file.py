import csv
from zipfile import ZipFile
from PyPDF2 import PdfReader
from openpyxl.reader.excel import load_workbook
from config import ZIP_DIR
from io import TextIOWrapper


def test_pdf_content(create_archive):
    with (ZipFile(ZIP_DIR) as zip_file):
        with zip_file.open("file.pdf") as pdf_file:
            reader = PdfReader(pdf_file)
            page = reader.pages[0]
            text = page.extract_text()

            assert "Тестовый PDF документ" in text

def test_xlsx_content(create_archive):
    with (ZipFile(ZIP_DIR) as zip_file):
        with zip_file.open("file.xlsx") as xlsx_file:
            workbook = load_workbook(xlsx_file)
            sheet = workbook.active

            assert sheet.max_row == 8
            assert sheet.cell(row=6, column=2).value == "Дмитрий"
            assert sheet.cell(row=2, column=3).value == 20

def test_csv_content(create_archive):
    with (ZipFile(ZIP_DIR) as zip_file):
        with zip_file.open("file.csv") as csv_file:
            csvreader = list(csv.reader(TextIOWrapper(csv_file, 'utf-8-sig')))
            rows = list(csvreader)

            assert len(rows) == 8
            assert rows[2] == ['12331', 'Анастасия', '28']